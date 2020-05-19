import openpyxl
from openpyxl.utils import FORMULAE
import openpyxl.styles


print('SUM' in FORMULAE)
# 求和公式在函数里,导入成功

wb=openpyxl.load_workbook('test.xlsx')
print(wb.sheetnames)
ws=wb['Sheet']

for row in ws.iter_rows(min_col=2,min_row=2,max_col=5,max_row=5):
    # row是每一行，row[3]就是每一行的第4列位置(注意是从min_col开始算的）
    ws[row[3].coordinate]='=SUM(%s:%s)'%(row[0].coordinate,row[2].coordinate)

center_alignment=openpyxl.styles.Alignment(horizontal='center',vertical='center')
for row in ws.iter_rows(min_col=2,min_row=2,max_col=6,max_row=5):
    ws[row[4].coordinate]='=IF(%s>250,"A","B")'%(row[3].coordinate)
    # Excel里只认双引号解析
    ws[row[4].coordinate].alignment=center_alignment


# LOOKUP函数查询（只能单行单列，切自动模糊匹配（没有确切值自动向下匹配）
ws['I2']='=LOOKUP(H2,D2:D5,A2:A5)'
ws['I2'].alignment=center_alignment
# 三个参数分别为查询值、查询范围、对应结果范围


# VLOOKUP函数
ws['J2']='=VLOOKUP(I2,A1:E5,5,FALSE)'
# 五个参数分别为查询值，查询范围，结果所在列（1开始数），是否进行模糊匹配
ws['J2'].alignment=center_alignment


wb.save('test.xlsx')