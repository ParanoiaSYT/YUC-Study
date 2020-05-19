import openpyxl
import datetime


wb=openpyxl.Workbook()

ws=wb.active
ws.append(['文本','数字'])

ws['A2']='520'
ws['B2']=520

ws['A3']=88.88
ws['A3'].number_format='#,###.00 $'

ws['B3']=datetime.datetime.today()
ws['B3'].number_format='yyyy-mm-dd'


# 设置数字格式
from openpyxl.styles.colors import RED,GREEN,BLUE,YELLOW
ws['A4'].number_format='[RED]+#,###.00;[GREEN]-#,###.00'
ws['A4']=999
ws['A5'].number_format='[RED]+#,###.00;[GREEN]-#,###.00'
ws['A5']=-999
# 通过分号为正负值做区分
ws['A6'].number_format='[RED][GREEN][BLUE][YELLOW]'
ws['A6']=0
ws['A7'].number_format='[RED];[GREEN];[BLUE];[YELLOW]'
ws['A7']='这是文本'
# 正值;负值;零值;文本
# 通过分号来设置格式


# 设置附加条件
ws['A8'].number_format='[=1]答对了;[=0]答错了'
ws['A8']=0
ws['A9'].number_format='[=1]答对了;[=0]答错了'
ws['A9']=1
ws['A10'].number_format='[=1]答对了;[=0]答错了'
ws['A10']=2
# 不符合条件的会乱码######

ws['A11'].number_format='[<60][RED]不及格;[>=60][BLUE]及格'
ws['A11']=59



wb.save('文本数字.xlsx')
