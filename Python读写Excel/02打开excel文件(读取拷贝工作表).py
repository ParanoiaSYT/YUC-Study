import openpyxl


wb=openpyxl.load_workbook('/Users/sunyuting/YuC-Study/爬虫项目/豆瓣top.xlsx')
# 打开excel文件(返回工作簿对象)
print(type(wb))


# 获取工作簿里所有工作表名字(返回一个列表类型）
print(wb.sheetnames)
print(type(wb.sheetnames))

ws=wb['Sheet']
# 获取指定名字的工作表(名字不存在会报错）

# 创建和删除工作表(默认创建在现有工作表之后）
nws=wb.create_sheet(index=0,title='FishC demo')
# 指定位置和名字
print(wb.sheetnames)

wb.remove(wb['FishC demo'])
# 删除FishC demo的工作表
print(wb.sheetnames)

# 定位单元格
c=ws['A2']
print(c.row,c.column,c.coordinate)
print(c.value)

# 距离A2两行的位置
d=c.offset(2,0)
print(d.coordinate,d.value)

# 获取指定列对应的字母（16进制）以及字母对应的列数
import openpyxl.utils
print(openpyxl.utils.get_column_letter(496))
print(openpyxl.utils.column_index_from_string('JB'))


# 访问多个单元格,(A2到B10范围)
print(type(ws['A2':'B10']))
for each_movie in ws['A2':'B10']:
# 工作表中的每一行是元组形式
    for each_cell in each_movie:
    # 每一行中的每一列是一个元组形式
        print(each_cell.value,end=' ')
    print('\n')
print(type(each_movie))


# 迭代所有行
for each_row in ws.rows:
    # 注意这里是rows
    print(each_row[0].value)
    # 打印每一行的第一个元素（电影名称）

print('=========================================')
# 迭代指定行
for each_row in ws.iter_rows(min_row=2,min_col=1,max_row=4,max_col=2):
    # 指定第二行第一列到第四行第二列
    print(each_row[0].value)


# 拷贝工作表
new=wb.copy_worksheet(ws)
new.title="copy version"
print(type(new))
wb.save('/Users/sunyuting/YuC-Study/爬虫项目/豆瓣top.xlsx')