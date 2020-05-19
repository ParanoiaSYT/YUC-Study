import openpyxl


wb=openpyxl.Workbook()
print(wb.sheetnames)
wb.remove(wb['Sheet'])

ws1=wb.create_sheet(title="小甲鱼",index=0)
ws2=wb.create_sheet(title="懒羊羊")
ws3=wb.create_sheet(title="皮鞋")
ws4=wb.create_sheet(title="飞燕")
# 个性化工作表标签栏
ws1.sheet_properties.tabColor="FF0000"
ws2.sheet_properties.tabColor="66CCFF"
ws3.sheet_properties.tabColor="8B008B"
ws4.sheet_properties.tabColor="0000FF"


# 调整行高和列宽
ws2.row_dimensions[2].height=100
ws2.column_dimensions['C'].width=50
# 注意虽然宽50，高100，但由于高、宽单位不同，所以显示不一样


# 合并拆分单元格
ws1.merge_cells("A1:C3")
ws1['A1']="I love fishc.com"
# 合并后只能赋值给左上角的位置（代表合并后整个单元格）

ws1.unmerge_cells("A1:C3")
# 拆分的范围要与合并的范围统一
# 拆分后留下的字在A1（左上角位置）

wb.save("FishCC.xlsx")
# 每次运行会覆盖保存

# 冻结窗口
# 打开现有的excel文件
wb=openpyxl.load_workbook("/Users/sunyuting/YuC-Study/爬虫项目/豆瓣top.xlsx")
ws=wb.active
ws.freeze_panes='C2'
# 表示C2位置左边和上面的单元格都被冻结，移动到别的单元格永远会显示

# ws.freeze_panes='A1'
ws.freeze_panes=None
# 解冻



wb.save('/Users/sunyuting/YuC-Study/爬虫项目/豆瓣top.xlsx')


