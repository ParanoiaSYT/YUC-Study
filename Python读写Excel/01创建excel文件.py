import openpyxl
import datetime



wb=openpyxl.Workbook()
# 实例化Workbook类，创建工作簿

ws=wb.active
# 获取工作簿中的一个工作表
print(ws.title)

ws['A1']=520
# 加到A1位置
ws.append([2,3,5])
# 默认加到接下来的A2，B2，C2位置
ws['A3']=datetime.datetime.now()
# 还可以存放python类型的值

wb.save('demo.xlsx')
# 保存excel格式文件


