import openpyxl
import openpyxl.styles

wb=openpyxl.Workbook()
ws=wb.active

b2=ws['B2']
b2.value="fishc"
bold_red_font=openpyxl.styles.Font(bold=True,color="FF0000")
b2.font=bold_red_font

b3=ws['B3']
b3.value="CCCC"
italic_strike_blue_16font=openpyxl.styles.Font(size=16,italic=True,strike=True,color="0000FF")
# italic倾斜,strike删除线
b3.font=italic_strike_blue_16font


# 填充单元格
from openpyxl.styles import PatternFill
# 纯色填充
yellow_fill=PatternFill(fill_type='solid',fgColor="FFFF00")
b2.fill=yellow_fill

from openpyxl.styles import GradientFill
# 渐进填充
red2green_fill=GradientFill(type='linear',stop=("FF0000","00FF00"))
# 填充方式线性，填充样式是一个元组形式（红-绿）
b3.fill=red2green_fill

# 绘制边框
from openpyxl.styles import Border,Side
thin_side=Side(border_style="thin",color="000000")
double_side=Side(border_style="double",color="0000FF")
# 绘制对角线
b2.border=Border(diagonal=thin_side,diagonalUp=True,diagonalDown=True)
# 绘制上下左右边框
b3.border=Border(left=double_side,top=double_side,right=double_side,bottom=double_side)


# 文本对齐
ws.merge_cells('A4:C5')
ws['A4']='I love fishc'
centr_alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
# 水平垂直居中
ws['A4'].alignment=centr_alignment


# 命名样式
heighlight=openpyxl.styles.NamedStyle(name="heighlight")
heighlight.font=openpyxl.styles.Font(bold=True,size=20)
heighlight.alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
# 命名了一个叫heighlight的样式模版
# 注册到工作簿里
wb.add_named_style(heighlight)

ws['A4'].style=heighlight
ws['D5']="我自横刀向天笑"
ws['D5'].style=heighlight

wb.save('样式.xlsx')
