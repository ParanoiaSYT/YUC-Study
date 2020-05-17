import requests
from bs4 import BeautifulSoup
import openpyxl



def get_movies():

    headers={'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.122 Safari/537.36'}
    movie_list=[]

    for i in range(10):

        link = 'http://movie.douban.com/top250?'+'start=%d'%(i*25)
        r=requests.get(link,headers=headers,timeout=10)

        print(str(i+1),'页状态响应码：',r.status_code)
        # print(r.text)

        soup=BeautifulSoup(r.text,'lxml')
        title_list=soup.find_all('div',class_='hd')
        rank_list=soup.find_all('span',class_='rating_num')
        descibe_list=soup.find_all('span',class_='inq')

        for j in range(len(title_list)):
            title=title_list[j].a.span.text.strip()
            rank=rank_list[j].text.strip()
            try:
                describe=descibe_list[j].text.strip()
            except:
                describe='None'

            movie_list.append([title,rank,describe])
    return movie_list

movies=get_movies()
print(movies)


def save2excel(result):
    wb=openpyxl.Workbook()
    ws=wb.active

    ws['A1']='电影名称'
    ws['B1']='评分'
    ws['C1']='评语'
    for each in result:
        ws.append(each)
    wb.save("豆瓣top.xlsx")

save2excel(movies)